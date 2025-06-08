import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from email_handler import download_attachment_by_filename_or_subject
import streamlit as st # Import streamlit for st.info and st.error
import google.generativeai as genai # Import genai
import io # Import io for string operations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def download_latest_employee_schedule():
    """Downloads the latest employee schedule Excel file from email."""
    st.info("Attempting to download latest employee schedule...")
    filter_text = "ROSATI'S EMPLOYEE SCHEDULE"
    download_path = download_attachment_by_filename_or_subject(filter_text, allowed_extensions=(".xlsx",))
    if download_path:
        st.success(f"Downloaded schedule: {os.path.basename(download_path)}")
    return download_path

def parse_employee_schedule(file_path):
    """Parses the employee schedule Excel file into a pandas DataFrame."""
    if not file_path or not os.path.exists(file_path):
        st.error("Schedule file not found.")
        return pd.DataFrame()

    try:
        full_df_raw = pd.read_excel(file_path, engine='openpyxl', header=None)

        header_row_index = -1
        days_of_week_keywords = ["MON", "TUES", "WED", "THURS", "FRI", "SAT", "SUN"]

        # Find the header row by looking for days of the week
        for r_idx in range(full_df_raw.shape[0]):
            row_values = [str(cell).strip().upper() for cell in full_df_raw.iloc[r_idx].dropna().tolist()]
            # Check if a significant number of day keywords are present in the row
            if sum(1 for day_keyword in days_of_week_keywords if day_keyword in row_values) >= 3:
                header_row_index = r_idx
                break
        
        if header_row_index == -1:
            st.error("Could not detect schedule header row. Ensure day names (Mon, Tue, etc.) are present.")
            return pd.DataFrame()
        
        # Extract column names
        columns = ["Employee Name"]
        day_name_standardization = {
            "MON": "Mon", "TUES": "Tue", "WED": "Wed", "THURS": "Thu",
            "FRI": "Fri", "SAT": "Sat", "SUN": "Sun"
        }
        for col_idx in range(1, full_df_raw.shape[1]): # Start from the second column
            cell_value = str(full_df_raw.iloc[header_row_index, col_idx]).strip().upper()
            if cell_value in day_name_standardization:
                columns.append(day_name_standardization[cell_value])
            else:
                # Stop adding columns if we encounter something that's not a day, or is empty
                if not cell_value:
                    break
                # If it's a non-day value and not empty, it might be a merged cell or other info, just add it for now.
                columns.append(f"Column_{col_idx}") 


        # Find the actual data start row
        data_start_row_index = -1
        for r_idx in range(header_row_index + 1, full_df_raw.shape[0]):
            first_col_value = str(full_df_raw.iloc[r_idx, 0]).strip()
            if first_col_value and not any(keyword in first_col_value.upper() for keyword in ["SERVERS:", "SUPPORT:", "WEEK OF", "ROSATI'S"]):
                data_start_row_index = r_idx
                break

        if data_start_row_index == -1:
            st.warning("No actual schedule data found after header. The schedule might be empty or formatted unexpectedly.")
            return pd.DataFrame()

        # Create DataFrame from data rows, using the dynamically extracted columns
        df = full_df_raw.iloc[data_start_row_index:].copy()
        df.columns = columns[:df.shape[1]] # Ensure column count matches data

        # Fill any NaN values in the schedule columns with empty strings
        schedule_columns = [col for col in columns if col != "Employee Name"]
        for col in schedule_columns:
            if col in df.columns:
                df[col] = df[col].fillna('')

        # Drop rows where 'Employee Name' is empty or contains section headers
        df = df[df["Employee Name"].astype(str).str.strip() != ""].copy()
        df = df[~df["Employee Name"].astype(str).str.upper().isin(["SERVERS:", "SUPPORT:"])].copy()
        df.dropna(how='all', inplace=True)

        if df.empty:
            st.warning("Parsed schedule is empty after processing.")
            return pd.DataFrame()

        st.success("Successfully parsed schedule file.")
        return df
    except Exception as e:
        st.error(f"Error parsing schedule file: {e}")
        return pd.DataFrame()

def generate_ai_schedule_changes(current_df: pd.DataFrame, user_prompt: str) -> pd.DataFrame:
    """
    Generates or adjusts employee schedule based on user prompt using Google's Gemini.
    """
    try:
        # Convert DataFrame to a string format for the prompt
        current_schedule_str = current_df.to_csv(index=False)

        prompt = f"""As a helpful AI assistant specializing in employee scheduling for a restaurant,
        you are given the current employee schedule in CSV format and a user's request for changes.

        Current Schedule (CSV):
        {current_schedule_str}

        User Request: "{user_prompt}"

        Please analyze the user's request in the context of the current schedule.
        Propose a revised schedule, strictly adhering to the CSV format provided for the current schedule.
        If a change is requested, make the change in the corresponding cell.
        If no change is explicitly requested for a particular employee or day, keep their existing schedule.
        Ensure all columns from the original schedule are present in the output.
        If a day-off is requested, put "OFF" in the corresponding cell.
        If a time change is requested, update the time in the corresponding cell.

        Output ONLY the revised schedule in CSV format, without any additional text or explanations.
        Make sure the header row is identical to the input CSV.
        """

        model = genai.GenerativeModel('gemini-1.5-pro-latest')
        response = model.generate_content(prompt)
        ai_response_text = response.text.strip()

        # Attempt to parse the AI's response back into a DataFrame
        try:
            updated_df = pd.read_csv(io.StringIO(ai_response_text))
            st.success("AI successfully generated a revised schedule!")
            return updated_df
        except Exception as parse_e:
            st.error(f"Failed to parse AI's response into a DataFrame: {parse_e}")
            st.warning("Please review the raw AI response above to understand the issue.")
            return current_df # Return original if parsing fails

    except Exception as e:
        st.error(f"Failed to generate AI schedule changes: {e}")
        return current_df # Return original DataFrame on error

def generate_formatted_excel_schedule(df: pd.DataFrame, file_path: str) -> BytesIO:
    """
    Generates an Excel file with the employee schedule in a polished format.
    """
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Schedule"

    # Read the raw Excel to get the dates for the header
    raw_excel_data = pd.read_excel(file_path, engine='openpyxl', header=None)

    # --- Title Section ---
    ws.merge_cells('A1:H2')
    title_cell = ws['A1']
    title_cell.value = "Rosati's Schedule"
    title_cell.font = Font(name='Arial', size=28, bold=True, color="FFFFFF") # Increased font size
    title_cell.fill = PatternFill(start_color="000033", end_color="000033", fill_type="solid") # Dark blue
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Headers for Schedule Data ---
    header_start_row = 4 # Adjusted row for headers as Department/Week Ending are gone
    # 'Employee and Assignment' header
    ws.merge_cells(start_row=header_start_row, end_row=header_start_row + 1, start_column=1, end_column=1)
    employee_header_cell = ws.cell(row=header_start_row, column=1, value="Employee and Assignment")
    employee_header_cell.alignment = Alignment(horizontal="center", vertical="center")
    employee_header_cell.font = Font(bold=True, size=14) # Increased font size

    # Days of the week headers (Mon-Sun)
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    current_col = 2

    # Find the row with the dates (row 7 in image, 0-indexed is 6)
    # The header_row_index from parse_employee_schedule is the row with 'Mon', 'Tue', etc.
    # The dates are in the row *after* this.
    date_row_for_display_index = 6 # This is the 0-indexed row number where actual dates (6/16, 6/17 etc.) are in the original Excel.

    extracted_dates = []
    for col_idx in range(1, len(days) + 1):
        # Get the cell value from the raw Excel data at the determined date_row_for_display_index
        date_value = raw_excel_data.iloc[date_row_for_display_index, col_idx]
        if pd.notna(date_value):
            try:
                # Convert to datetime and then format as M/D
                # Ensure date_value is a string before converting to datetime to handle various Excel cell types
                formatted_date = pd.to_datetime(str(date_value).split(' ')[0]).strftime('%#m/%#d') # Extract only date part
                extracted_dates.append(formatted_date)
            except Exception as e:
                # Fallback if date parsing fails
                st.warning(f"Could not parse date value {date_value}: {e}")
                extracted_dates.append('') # Append empty string if date parsing fails
        else:
            extracted_dates.append('')

    for idx, day in enumerate(days):
        ws.cell(row=header_start_row, column=current_col + idx, value=day).font = Font(bold=True, size=14) # Increased font size
        # Use extracted dates for the second header row
        date_to_display = extracted_dates[idx] if idx < len(extracted_dates) else ""
        ws.cell(row=header_start_row + 1, column=current_col + idx, value=date_to_display).font = Font(bold=True, size=14) # Increased font size
        ws.cell(row=header_start_row, column=current_col + idx).alignment = Alignment(horizontal="center")
        ws.cell(row=header_start_row + 1, column=current_col + idx).alignment = Alignment(horizontal="center")

    # Apply border to headers
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for r in range(header_start_row, header_start_row + 2):
        for c in range(1, len(days) + 2):
            ws.cell(row=r, column=c).border = thin_border

    # --- Populate Data ---
    data_start_row = header_start_row + 2
    for r_idx, row_data in df.iterrows():
        employee_name = row_data.get('Employee Name', '') # Assuming this column exists
        employee_cell = ws.cell(row=data_start_row + r_idx, column=1, value=employee_name)
        employee_cell.border = thin_border
        employee_cell.font = Font(size=12) # Increased font size

        for c_idx, day_name in enumerate(days):
            # Assuming days are columns in the DataFrame
            schedule_value = row_data.get(day_name, '') # Use .title() to match standardized columns, changed to `day_name` as columns should be standardized now
            # Ensure schedule_value is a string for Excel compatibility
            cell_value_str = str(schedule_value)
            if cell_value_str == '<NA>' or cell_value_str == 'nan':
                cell_value_str = ''
            cell = ws.cell(row=data_start_row + r_idx, column=c_idx + 2, value=cell_value_str)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=12) # Increased font size

    # Adjust column widths
    min_column_width = 10 # Minimum width for all columns
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column # Get the column number (1-indexed)
        for cell in column_cells:
            try:
                if cell.value is not None:
                    # Only consider non-merged cells or the top-left cell of a merged block for width calculation
                    if cell.coordinate not in ws.merged_cells:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            except:
                pass
        adjusted_width = max(min_column_width, (max_length + 2)) # Ensure minimum width
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Remove Footer (Copyright)
    # copyright_row = data_start_row + len(df) + 2
    # ws.merge_cells(start_row=copyright_row, end_row=copyright_row, start_column=1, end_column=len(days) + 1)
    # copyright_cell = ws.cell(row=copyright_row, column=1, value="© Copyright, 2014, Jaxworks, All Rights Reserved.")
    # copyright_cell.font = Font(color="0000FF", underline="single") # Blue and underlined
    # copyright_cell.alignment = Alignment(horizontal="center")


    wb.save(output)
    output.seek(0)
    return output