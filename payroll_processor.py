
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import tempfile

REFERENCE_DATA = {
    123: {"Name": "Krish Patel", "Job": "Cook", "Rate": 13.0, "Base Pay": 416.0},
    110: {"Name": "Sonu Mitha", "Job": "Manager", "Rate": 17.0, "Base Pay": 1020.0},
}

def process_payroll_file(uploaded_file):
    df = pd.read_excel(uploaded_file, skiprows=6)

    if "Employee ID" not in df.columns:
        raise ValueError("Missing 'Employee ID' column. Please make sure the Excel file has headers starting on row 7 with 'Employee ID' as one of the columns.")

    df = df[df["Employee ID"].notna()]
    df["Employee ID"] = df["Employee ID"].astype(int)

    df["Base Pay"] = df["Employee ID"].apply(lambda x: REFERENCE_DATA.get(x, {}).get("Base Pay", None))
    df["Rate"] = df["Employee ID"].apply(lambda x: REFERENCE_DATA.get(x, {}).get("Rate", None))
    df["Name"] = df["Employee ID"].apply(lambda x: REFERENCE_DATA.get(x, {}).get("Name", None))

    if "Other Tips" not in df.columns:
        raise ValueError("Missing 'Other Tips' column. Please verify the file includes this column.")

    df["Total Pay"] = df["Base Pay"] + df["Other Tips"]

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    df.to_excel(temp_file.name, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(temp_file.name)
    ws = wb.active

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    for row in range(2, ws.max_row + 1):
        ws[f"O{row}"] = f"=N{row}+F{row}"

    wb.save(temp_file.name)
    return df, temp_file.name
