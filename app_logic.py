import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st # Streamlit is used for debugging messages, adjust if not desired in app_logic
import os
import json
from pathlib import Path

# Define the path for the rates JSON file
RATES_FILE = Path("rates.json")

def normalize_name(name: str) -> str:
    """Normalizes an employee name for consistent comparison (lowercase, no commas, single spaces)."""
    if not isinstance(name, str):
        return ""
    name = name.lower().replace(',', '')
    return ' '.join(name.split())

# Employees who never need a rate prompt, regardless of missing rate/hours
# These names will be normalized for internal comparison
RAW_EXCLUDED_EMPLOYEES_FROM_RATE_PROMPT = ["Kush Patel", "Krish Patel", "Sonu Mitha", "A, Angie", "delivery delivery driver", "jayesh"]
EXCLUDED_EMPLOYEES_FROM_RATE_PROMPT = [normalize_name(name) for name in RAW_EXCLUDED_EMPLOYEES_FROM_RATE_PROMPT]

def load_reference_rates():
    """Load reference rates from rates.json or initialize if not exists."""
    if RATES_FILE.exists():
        with open(RATES_FILE, 'r') as f:
            # Ensure keys are normalized on load for string names, and IDs are ints
            loaded_rates = json.load(f)
            normalized_rates = {}
            for k, v in loaded_rates.items():
                if isinstance(k, str) and not k.isdigit():
                    normalized_rates[normalize_name(k)] = v
                else:
                    normalized_rates[int(k)] = v
            return normalized_rates
    else:
        # Initial hardcoded rates if file doesn't exist
        initial_rates = {
            44: 15.0, 130: 9.0, 71: 10.0, 74: 10.5, 136: 9.0, 117: 15.0, 110: 17.0, 79: 15.0, 123: 13.0,
            112: 23.0, 135: 9.0, 12: 11.0, 15: 15.0, 140: 9.0, 11: 15.0, 141: 9.0, 143: 9.0, 144: 9.0,
            145: 9.0, 142: 9.0, 146: 9.0, 147: 9.0 # Added Witt, Lacie's rate
        }
        # Also add names to initial rates, normalized
        initial_rates["Kush Patel"] = 0.0 # Kush Patel has 0.0 base pay
        initial_rates["Krish Patel"] = 10.40 # Krish Patel hardcoded
        initial_rates["Sonu Mitha"] = 15.00 # Sonu Mitha hardcoded
        initial_rates["A, Angie"] = 0.0 # Example: if Angie is salaried or doesn't have a rate
        initial_rates["delivery delivery driver"] = 0.0 # Example
        initial_rates["jayesh"] = 0.0 # Example

        # Save normalized string names to JSON for consistency
        json_serializable_rates = {str(k): v for k, v in initial_rates.items()}

        with open(RATES_FILE, 'w') as f:
            json.dump(json_serializable_rates, f, indent=4)
        return initial_rates

def save_reference_rates(rates):
    """Save updated reference rates to rates.json."""
    # Convert integer keys to string for JSON serialization
    json_serializable_rates = {str(k): v for k, v in rates.items()}
    with open(RATES_FILE, 'w') as f:
        json.dump(json_serializable_rates, f, indent=4)

# Load rates at module startup
reference_rates = load_reference_rates()

def process_payroll_excel(file_path):
    try:
        if not file_path or not os.path.exists(file_path):
            st.error(f"Error: File not found or invalid path: {file_path}")
            return pd.DataFrame()

        st.write(f"DEBUG: Attempting to read Excel file. Path: {file_path}, Type: {type(file_path)}")
        # Step 1: Read the Excel file without a header to dynamically find the header row
        # Pass the file_path directly; pandas handles opening and closing
        df_raw_initial = pd.read_excel(file_path, header=None)

        header_row_index = None
        # Standardized expected headers for the OUTPUT DataFrame columns
        expected_headers = ["ID", "Name", "Job Description", "Rate", "Hours", "Base Pay", "Driver Reim.", "CC Tips", "Cash Tips", "Other Tips", "Total Pay"]

        # Search for the header row in the first few rows (e.g., first 20 rows to be safe)
        header_search_keywords = ["ID", "Name", "Base Pay", "Total Pay"]
        for i in range(min(len(df_raw_initial), 20)):
            current_row_str = [str(x).strip() for x in df_raw_initial.iloc[i].tolist()] # Strip whitespace

            # Check if all critical headers are present in this row
            if all(keyword in current_row_str for keyword in header_search_keywords):
                header_row_index = i
                break

        if header_row_index is None:
            st.error(f"Could not find a row containing essential headers ({', '.join(header_search_keywords)}). Please check the Excel file format.")
            return pd.DataFrame()

        # Step 2: Read the Excel file again, this time with the detected header row
        df_raw = pd.read_excel(file_path, header=header_row_index)

        print(f"DEBUG: Columns after initial read: {df_raw.columns.tolist()}")

        # Normalize column names for easier access (e.g., remove newlines, extra spaces)
        # Map detected columns to internal, clean names
        column_mapping_input_to_internal = {
            "ID": "ID",
            "Name": "Name",
            "Job Desc": "Job_Desc",
            "Rate": "Rate",
            "Hours": "Hours",
            "Base Pay": "Base_Pay_Excel",
            "Driver\nReim": "Driver_Reim",  # Handle newline in column name
            "CC/\nOther Tips": "CC_Tips_Raw",  # Handle newline in column name
            "Cash\nTips": "Cash_Tips_Raw",  # Handle newline in column name
            "Total Tips": "Total_Tips_Raw",
            "Subtotal": "Subtotal",
            "Meal Accts": "Meal_Accts",
            "Total Pay": "Total_Pay_Excel"
        }

        # Rename columns in df_raw for easier and consistent access
        df_raw.rename(columns=column_mapping_input_to_internal, inplace=True)

        # Drop any 'Unnamed' columns if they exist and are not needed
        df_raw = df_raw.loc[:, ~df_raw.columns.str.contains('^Unnamed')]

        # Print the first few rows of the processed DataFrame for debugging
        print("\nDEBUG: First few rows of processed DataFrame:")
        print(df_raw.head().to_string())

        # Standardized expected headers for the OUTPUT DataFrame columns
        expected_output_headers = ["ID", "Name", "Job Description", "Rate", "Hours", "Base Pay",
                                   "Driver Reim.", "CC Tips", "Cash Tips", "Other Tips", "Total Pay"]
        df_output = pd.DataFrame(columns=expected_output_headers).copy()

        processed_rows = []
        i = 0
        while i < len(df_raw):
            current_row = df_raw.iloc[i]

            # Primary row identification: non-empty ID and Name
            id_raw = current_row.get("ID")
            name_raw = current_row.get("Name")

            if pd.isna(id_raw) or pd.isna(name_raw):
                i += 1
                continue

            try:
                id_val = int(id_raw)
                name = str(name_raw).strip() if pd.notna(name_raw) else ""
                normalized_name = normalize_name(name) # Normalize the name for comparison

                # Initialize values from current_row (main row)
                cc_tips = float(current_row.get("CC_Tips_Raw", 0.0)) if pd.notna(current_row.get("CC_Tips_Raw")) else 0.0
                cash_tips = float(current_row.get("Cash_Tips_Raw", 0.0)) if pd.notna(current_row.get("Cash_Tips_Raw")) else 0.0
                driver_reim = float(current_row.get("Driver_Reim", 0.0)) if pd.notna(current_row.get("Driver_Reim")) else 0.0
                base_pay_from_excel = float(current_row.get("Base_Pay_Excel", 0.0)) if pd.notna(current_row.get("Base_Pay_Excel")) else 0.0
                total_pay_from_excel = float(current_row.get("Total_Pay_Excel", 0.0)) if pd.notna(current_row.get("Total_Pay_Excel")) else 0.0

                # Initialize job_desc, rate, hours (might be empty/NaN in current_row)
                job_desc = str(current_row.get("Job_Desc", "")).strip() if pd.notna(current_row.get("Job_Desc")) else ""
                
                # Get rate from reference_rates dictionary (check by ID then normalized name)
                rate = reference_rates.get(id_val, reference_rates.get(normalized_name, 0.0))
                hours = float(current_row.get("Hours", 0.0)) if pd.notna(current_row.get("Hours")) else 0.0

                # Check the next row for Job Desc, Rate, Hours, Driver Reim, CC/Other Tips, Cash Tips if not found in current row
                if i + 1 < len(df_raw) and pd.isna(df_raw.iloc[i+1].get("ID")):  # Check if next row is a detail row (no ID)
                    next_row = df_raw.iloc[i+1]

                    # Update fields from next_row ONLY if they are not already set in current_row or are NaN
                    job_desc = str(next_row.get("Job_Desc", job_desc)).strip() if pd.notna(next_row.get("Job_Desc")) else job_desc
                    
                    # Keep the rate from reference_rates dictionary, but ensure it's numeric
                    rate_from_next_row = next_row.get("Rate")
                    if pd.notna(rate_from_next_row):
                        try:
                            rate = float(rate_from_next_row) # Update rate if found in next row and is numeric
                        except (ValueError, TypeError):
                            rate = 0.0 # Explicitly set to 0.0 if conversion fails

                    hours = float(next_row.get("Hours", hours)) if pd.notna(next_row.get("Hours")) else 0.0 # Ensure hours from next row are numeric
                    driver_reim = float(next_row.get("Driver_Reim", driver_reim)) if pd.notna(next_row.get("Driver_Reim")) else driver_reim
                    cc_tips = float(next_row.get("CC_Tips_Raw", cc_tips)) if pd.notna(current_row.get("CC_Tips_Raw")) else cc_tips
                    cash_tips = float(current_row.get("Cash_Tips_Raw", cash_tips)) if pd.notna(current_row.get("Cash_Tips_Raw")) else cash_tips
                    i += 1  # Consume the next row as part of the current record

                # Base Pay Calculation Logic
                base_pay = 0.0  # Initialize to 0 for clear calculation flow

                # Auto-assign rates based on Job Description if rate is 0 and not an excluded employee
                if rate == 0 and normalized_name not in EXCLUDED_EMPLOYEES_FROM_RATE_PROMPT:
                    if "support" in job_desc.lower():
                        rate = 15.0
                        reference_rates[normalized_name] = rate # Store for future
                        save_reference_rates(reference_rates)
                    elif "server" in job_desc.lower():
                        rate = 9.0
                        reference_rates[normalized_name] = rate # Store for future
                        save_reference_rates(reference_rates)

                # Hardcoded rates and hours for specific employees
                if id_val == 123:  # Krish Patel
                    rate = 10.40 # Hardcoded rate for Krish
                    hours = 40.0 # Hardcoded hours for Krish
                    base_pay = hours * rate
                elif id_val == 110:  # Sonu Mitha
                    rate = 15.00 # Hardcoded rate for Sonu
                    hours = 68.0 # Hardcoded hours for Sonu
                    base_pay = hours * rate
                elif id_val == 4:  # Kush Patel
                    base_pay = 0.0
                elif rate > 0 and hours > 0:
                    base_pay = hours * rate
                elif pd.notna(base_pay_from_excel) and base_pay_from_excel > 0:
                    base_pay = base_pay_from_excel
                elif pd.notna(total_pay_from_excel) and total_pay_from_excel > 0 and (total_pay_from_excel >= (cc_tips + cash_tips + driver_reim)):
                    base_pay = total_pay_from_excel - cc_tips - cash_tips - driver_reim
                else:
                    base_pay = 0.0

                # Calculate hours by backsolving: hours = base_pay / rate
                # This line will be implicitly handled for Krish and Sonu as their hours are hardcoded
                hours_calc = hours # Use the hardcoded hours for Krish/Sonu, or calculated hours otherwise
                if rate > 0 and normalized_name not in EXCLUDED_EMPLOYEES_FROM_RATE_PROMPT: # Only back-solve for others, excluding specified employees
                    hours_calc = base_pay / rate

                other_tips = cc_tips + cash_tips
                total_pay_calculated = base_pay + other_tips

                # Print final calculated values for debugging
                print(f"  Extracted: Rate={rate}, Hours={hours}, Driver Reim={driver_reim}, CC Tips={cc_tips}, Cash Tips={cash_tips}")
                print(f"  Calculated: Base Pay={base_pay}, Other Tips={other_tips}, Total Pay={total_pay_calculated}")
                print(f"  Backsolved Hours: {hours_calc}")

                processed_rows.append({
                    "ID": id_val,
                    "Name": name,
                    "Job Description": job_desc,
                    "Rate": rate,
                    "Hours": round(hours_calc, 2),
                    "Base Pay": round(base_pay, 2),
                    "Driver Reim.": round(driver_reim, 2),
                    "CC Tips": round(cc_tips, 2),
                    "Cash Tips": round(cash_tips, 2),
                    "Other Tips": round(other_tips, 2),
                    "Total Pay": round(total_pay_calculated, 2)
                })

            except Exception as e:
                print(f"Error processing row {i} (ID: {id_raw}, Name: {name_raw}): {str(e)}")
                i += 1
                continue
            i += 1

        df_output = pd.DataFrame(processed_rows)

        if df_output.empty:
            print("No valid data was processed after dynamic header detection. Please check the Excel file format.")
        else:
            # Calculate totals row
            total_base_pay = df_output["Base Pay"].sum()

            print(f"Successfully processed {len(df_output)} rows of data.")

        return df_output

    except Exception as e:
        print(f"Critical error during Excel file processing: {str(e)}")
        return pd.DataFrame() # Return empty DataFrame on error

def generate_excel_download(df, filename="Final_Payroll_Report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(r)
        current_row = ws[r_idx]
        for cell in current_row:
            cell.font = Font(name='Arial', size=11)
            cell.alignment = Alignment(horizontal='left')
            
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            
            medium_bottom_border = Border(left=Side(style='thin'),
                                          right=Side(style='thin'),
                                          top=Side(style='thin'),
                                          bottom=Side(style='medium'))
            
            medium_top_border = Border(left=Side(style='thin'),
                                       right=Side(style='thin'),
                                       top=Side(style='medium'),
                                       bottom=Side(style='thin'))

            if r_idx == 1:  # Header row
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.border = medium_bottom_border
            elif r_idx == ws.max_row: # Totals row
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.border = medium_top_border
            else: # Data rows - apply no specific border/bolding, just default Arial 11 and left alignment
                cell.font = Font(name='Arial', size=11)
                cell.border = thin_border

    # Add Excel formula for Total Pay column dynamically based on final dataframe structure
    headers = [cell.value for cell in ws[1]] # Get headers from the first row of the worksheet
    base_pay_col_letter = None
    other_tips_col_letter = None
    total_pay_col_letter = None

    for col_idx, header_name in enumerate(headers, 1):
        if header_name == "Base Pay":
            base_pay_col_letter = chr(64 + col_idx)
        elif header_name == "Other Tips":
            other_tips_col_letter = chr(64 + col_idx)
        elif header_name == "Total Pay":
            total_pay_col_letter = chr(64 + col_idx)

    if base_pay_col_letter and other_tips_col_letter and total_pay_col_letter:
        for row_idx in range(2, ws.max_row):
            ws[f'{total_pay_col_letter}{row_idx}'] = f'={base_pay_col_letter}{row_idx}+{other_tips_col_letter}{row_idx}'

    # Autofit columns based on content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) # Add a small padding
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def process_payroll_report(file_path):
    df_final = process_payroll_excel(file_path)
    if not df_final.empty:
        excel_output = generate_excel_download(df_final)
        output_filename = "Final_Payroll_Report.xlsx"
        with open(output_filename, "wb") as f:
            f.write(excel_output.getvalue())
        return df_final, output_filename
    return pd.DataFrame(), None 